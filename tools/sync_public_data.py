#!/usr/bin/env python3
"""
Sync MUEC public website data from the private master spreadsheets.

WHY THIS EXISTS
  Some master spreadsheets contain data that must NOT go on the public website:
    - editor emails (Editorial Histories -> "Email" column)
    - unverified integrity allegations (Journal Profiles -> "Integrity note" column)
    - the entire Editors' Contacts and ISME Panel Responses files
  This script copies ONLY the public-safe spreadsheets (with sensitive columns
  removed) from the private folder into website/data/xlsx/, which is what the
  website reads and what gets published to GitHub Pages.

HOW TO USE
  1. Edit the master files in:  Editorial Data (PRIVATE - do not publish)/
     (this folder sits NEXT TO the website/ repo, one level up — never inside it)
  2. Run:  python3 website/tools/sync_public_data.py
  3. Commit & push the website/ folder (GitHub Desktop).

This script lives inside the website/ repo (tools/) so it is version-controlled,
but the private master folder is OUTSIDE the repo, so it is never published.
"""
import os, sys, shutil
try:
    import openpyxl
except ImportError:
    sys.exit("Please install openpyxl first:  pip3 install openpyxl")

# This file lives at  website/tools/sync_public_data.py
HERE    = os.path.dirname(os.path.abspath(__file__))   # .../website/tools
WEBSITE = os.path.dirname(HERE)                         # .../website
PROJECT = os.path.dirname(WEBSITE)                      # project root (holds the private folder)
PRIV = os.path.join(PROJECT, "Editorial Data (PRIVATE - do not publish)")
PUB  = os.path.join(WEBSITE, "data", "xlsx")
os.makedirs(PUB, exist_ok=True)

import re
_EMAIL = re.compile(r'[\w.+-]+@[\w.-]+\.\w+')
# An email is a publishable ROLE/INBOX address if its local-part contains one of these.
# Anything else (personal/individual emails) is redacted. Redact-by-default = safe on updates:
# an unrecognised address is hidden rather than leaked. To keep a new inbox, add its token here.
_ROLE_TOKENS = ["editor","journal","jrnl","submiss","publication","research","office","info",
                "contact","admin","society","revista","music","eloido",
                "jrme","bjme","crme","fjme","mjm","ijea","ijrcs","jpme","riem","abem","leeme",
                "cca","kmes","asme","onkyoiku","yyyj","zgyyjy","ijme","mej","pmer","vrme"]

def _redact_personal_emails(text):
    if not text: return text
    def repl(m):
        local = m.group(0).split("@")[0].lower()
        return m.group(0) if any(tok in local for tok in _ROLE_TOKENS) else "[editor email withheld]"
    return _EMAIL.sub(repl, str(text))

# --- Canonical Chinese journal names: "English (简体)" ----------------------------
# The "China journals" sheet in Journal_Profiles_Indexing.xlsx is the source of truth
# for simplified-character names ("《简体》 · English"). We normalise every Chinese
# journal name across the public data to "English (简体)" so they (a) sort
# alphabetically alongside the others and (b) read consistently — no pinyin, no
# Chinese-first. To add a journal, just add its row to that China-journals sheet.
_CANON = None
def _chinese_canon():
    global _CANON
    if _CANON is None:
        _CANON = {}
        wb = openpyxl.load_workbook(os.path.join(PRIV, "Journal_Profiles_Indexing.xlsx"))
        sh = "China journals (CSSCI · 北大核心)"
        if sh in wb.sheetnames:
            ws = wb[sh]; hdr = [c.value for c in ws[1]]
            jc = hdr.index("Journal (中文 · English)")
            for row in ws.iter_rows(min_row=2):
                m = re.match(r'^《(.+?)》\s*[·•]\s*(.+)$', str(row[jc].value or "").strip())
                if m:
                    simp, eng = m.group(1).strip(), m.group(2).strip()
                    _CANON[eng.lower()] = f"{eng} ({simp})"
    return _CANON

def canon_journal(raw):
    """Return a Chinese journal name as 'English (简体)'; pass other names through.
    Matches on the English portion, so it handles 'English / Pinyin',
    'English / Pinyin (简体)', and '《简体》 · English' inputs. Longest match wins."""
    if not raw: return raw
    s = str(raw); low = s.lower(); best = None
    for eng_l, canonical in _chinese_canon().items():
        if eng_l in low and (best is None or len(eng_l) > len(best[0])):
            best = (eng_l, canonical)
    return best[1] if best else s

def drop_column(path_in, path_out, sheet_name, column_header):
    wb = openpyxl.load_workbook(path_in)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        if column_header in headers:
            ws.delete_cols(headers.index(column_header) + 1, 1)
            print(f"   removed '{column_header}' from sheet '{sheet_name}'")
    wb.save(path_out)

def process_histories(path_in, path_out):
    """Replace the 'Email' column with a 'Status' column that keeps ONLY deceased
    ('in memoriam') notes — all real emails and email-availability notes are dropped."""
    wb = openpyxl.load_workbook(path_in)
    ws = wb["Editorial Histories"]
    headers = [c.value for c in ws[1]]
    if "Email" in headers:
        ci = headers.index("Email") + 1
        ws.cell(row=1, column=ci).value = "Status"
        for row in ws.iter_rows(min_row=2):
            cell = row[ci-1]
            s = str(cell.value or "")
            m = re.search(r'deceased[^/;]*', s, re.I)
            cell.value = m.group(0).strip(" ()") if m else ""
            # normalise "deceased (2001)" form
            if cell.value:
                yr = re.search(r'(\d{4})', cell.value)
                cell.value = "In memoriam" + (f" ({yr.group(1)})" if yr else "")
    # Normalise Chinese journal names to "English (简体)" (drops pinyin, English-first)
    if "Journal" in headers:
        jc = headers.index("Journal") + 1
        for row in ws.iter_rows(min_row=2):
            row[jc-1].value = canon_journal(row[jc-1].value)
    wb.save(path_out)
    print("   converted 'Email' -> 'Status' (in-memoriam only; emails dropped); normalised Chinese names")

def process_indexing(path_in, path_out):
    """Build ONE merged 'Profiles & Indexing' sheet: the Chinese journals are folded in as rows
    with a dedicated 'CSSCI / 北大核心' column (no separate China table). Integrity-note dropped,
    personal emails redacted."""
    from openpyxl import Workbook
    wb = openpyxl.load_workbook(path_in)
    main = [[c.value for c in row] for row in wb["Profiles & Indexing"].iter_rows()]
    headers = [str(h) if h is not None else "" for h in main[0]]
    body = main[1:]
    CSSCI = "CSSCI / 北大核心"
    notes_idx = headers.index("Notes") if "Notes" in headers else len(headers)
    headers.insert(notes_idx, CSSCI)
    body = [list(r[:notes_idx]) + [""] + list(r[notes_idx:]) for r in body]

    if "Journal contact / submissions" in headers:
        ci = headers.index("Journal contact / submissions")
        for r in body:
            if ci < len(r): r[ci] = _redact_personal_emails(r[ci])

    # Normalise Chinese names in the main list, and track which journals are present
    # so a journal that also appears in the China sheet is not duplicated.
    jci = headers.index("Journal") if "Journal" in headers else 0
    for r in body:
        if jci < len(r): r[jci] = canon_journal(r[jci])
    present = {str(r[jci]).strip().lower() for r in body if jci < len(r)}

    cn = "China journals (CSSCI · 北大核心)"
    if cn in wb.sheetnames:
        crows = [[c.value for c in row] for row in wb[cn].iter_rows()]
        ch = [str(h) if h is not None else "" for h in crows[0]]
        cv = lambda row, name: (row[ch.index(name)] if name in ch and ch.index(name) < len(row) else "") or ""
        for row in crows[1:]:
            if not any(v not in (None, "") for v in row): continue
            parts = []
            if str(cv(row, "CSSCI")).lower().startswith("y"): parts.append("CSSCI")
            if str(cv(row, "北大核心 (Peking Core)")).lower().startswith("y"): parts.append("北大核心")
            cssci_val = "; ".join(parts) if parts else str(cv(row, "CSSCI"))
            notes = "; ".join(x for x in [
                (f"Tier: {cv(row,'Tier')}" if cv(row,'Tier') else ""),
                (f"Music-ed column: {cv(row,'Music-education column?')}" if cv(row,'Music-education column?') else "")] if x)
            jname = canon_journal(cv(row, "Journal (中文 · English)"))
            if jname.strip().lower() in present: continue   # already in the main list
            present.add(jname.strip().lower())
            merged = {"Journal": jname, "ISSN (print / online)": "—",
                      "Publisher / type": cv(row, "Publisher / institution"), "Journal contact / submissions": "",
                      "Scopus": "No", "Web of Science": "No", "ERIC": "No", "RILM": "—",
                      "DOAJ / open access": "No", "Regional / other index": "CNKI", CSSCI: cssci_val, "Notes": notes}
            body.append([merged.get(h, "") for h in headers])

    out = Workbook(); ws = out.active; ws.title = "Profiles & Indexing"
    ws.append(headers)
    for r in body: ws.append(r)
    out.save(path_out)
    print(f"   merged China journals into one indexing table ({len(body)} rows, +'{CSSCI}' column)")

def copy_verbatim(path_in, path_out):
    shutil.copy2(path_in, path_out)
    print("   copied as-is")

JOBS = [
    # (master filename, published filename, action)
    ("Music_Ed_Journal_Editorial_Histories.xlsx", "editorial_histories.xlsx", process_histories),
    ("Journal_Profiles_Indexing.xlsx", "journal_indexing.xlsx", process_indexing),
    ("Editorial_Board_Archive_Pointers.xlsx", "board_archive_pointers.xlsx", copy_verbatim),
    ("Research_Integrity_Tracker.xlsx", "research_integrity.xlsx", copy_verbatim),
]

# Files that must NEVER be published (left in the private folder only):
PRIVATE_ONLY = ["Music_Ed_Journal_Editors_Contacts.xlsx", "ISME_2026_Panel_Responses.xlsx"]

def main():
    if not os.path.isdir(PRIV):
        sys.exit(f"Private folder not found: {PRIV}")
    for src, dst, action in JOBS:
        ip = os.path.join(PRIV, src)
        op = os.path.join(PUB, dst)
        if not os.path.exists(ip):
            print(f"!! missing master: {src} (skipped)"); continue
        print(f"-> {src}  ->  website/data/xlsx/{dst}")
        action(ip, op)
    print("\nNOT published (private):", ", ".join(PRIVATE_ONLY))
    print("Done. Now commit & push the website/ folder.")

if __name__ == "__main__":
    main()
